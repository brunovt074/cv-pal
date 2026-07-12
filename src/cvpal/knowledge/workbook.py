"""Builds data/cv-knowledge-base.xlsx - the MVP deliverable and the git-
versioned source of truth for all downstream agents (voice-style,
cv-generation, freelance-profiles).

See skills/data-analytics/SKILL.md for the schema rationale.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from cvpal.analytics.models import (
    CoverLetterEntry,
    CVVersionEntry,
    StructuredKnowledgeBase,
)

_VOICE_PROFILE_PLACEHOLDER_COLUMNS = ["trait", "value", "notes"]


def _cell_value(value: object) -> object:
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return value


def _write_model_sheet(wb: Workbook, name: str, rows: list[BaseModel]) -> Worksheet:
    ws = wb.create_sheet(name)
    if not rows:
        return ws

    fields = list(type(rows[0]).model_fields.keys())
    ws.append(fields)
    for row in rows:
        dumped = row.model_dump()
        ws.append([_cell_value(dumped[f]) for f in fields])
    return ws


def build_workbook(
    knowledge_base: StructuredKnowledgeBase,
    cv_versions: list[CVVersionEntry],
    cover_letters: list[CoverLetterEntry],
    output_path: Path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty "Sheet"

    _write_model_sheet(wb, "PersonalData", knowledge_base.personal_data)
    _write_model_sheet(wb, "Summary", knowledge_base.summaries)
    _write_model_sheet(wb, "Experience", knowledge_base.experience)
    _write_model_sheet(wb, "Education", knowledge_base.education)
    _write_model_sheet(wb, "Certifications", knowledge_base.certifications)
    _write_model_sheet(wb, "Skills", knowledge_base.skills)
    _write_model_sheet(wb, "Projects", knowledge_base.projects)
    _write_model_sheet(wb, "Languages", knowledge_base.languages)
    _write_model_sheet(wb, "CVVersions", cv_versions)
    _write_model_sheet(wb, "CoverLetters", cover_letters)

    # Populated later by `cvpal voice` (Fase 3) - reserve the tab now so the
    # schema is stable for anyone opening the workbook early.
    voice_ws = wb.create_sheet("VoiceProfile")
    voice_ws.append(_VOICE_PROFILE_PLACEHOLDER_COLUMNS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
