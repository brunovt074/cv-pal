import openpyxl

from cvpal.domain.knowledge.models import (
    CoverLetterEntry,
    CVVersionEntry,
    ExperienceBullet,
    KnowledgeBase,
)
from cvpal.infrastructure.persistence.xlsx_exporter import export_workbook


def test_export_workbook_creates_all_expected_tabs(tmp_path):
    kb = KnowledgeBase(
        experience=[
            ExperienceBullet(
                company="Acme",
                role="Dev",
                start_date="2020",
                end_date="Present",
                bullet="Did a thing",
                tech_tags=["Java", "Spring"],
            )
        ]
    )
    cv_versions = [
        CVVersionEntry(source_file="a.pdf", stack="Java", source_language="en", is_most_recent_variant=True)
    ]
    cover_letters = [
        CoverLetterEntry(source_file="b.pdf", source_language="es", original_text="hola", english_text="hello")
    ]

    output = tmp_path / "kb.xlsx"
    export_workbook(kb, cv_versions, cover_letters, output)

    assert output.exists()
    wb = openpyxl.load_workbook(output)
    expected_tabs = {
        "PersonalData", "Summary", "Experience", "Education", "Certifications",
        "Skills", "Projects", "Languages", "CVVersions", "CoverLetters", "VoiceProfile",
    }
    assert set(wb.sheetnames) == expected_tabs


def test_export_workbook_lists_become_comma_joined_strings(tmp_path):
    kb = KnowledgeBase(
        experience=[
            ExperienceBullet(
                company="Acme", role="Dev", start_date="2020", end_date="Present",
                bullet="Did a thing", tech_tags=["Java", "Spring"],
            )
        ]
    )
    output = tmp_path / "kb.xlsx"
    export_workbook(kb, [], [], output)

    wb = openpyxl.load_workbook(output)
    ws = wb["Experience"]
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    tech_tags_value = row[header.index("tech_tags")]
    assert tech_tags_value == "Java, Spring"


def test_export_workbook_handles_empty_lists_without_error(tmp_path):
    output = tmp_path / "empty.xlsx"
    export_workbook(KnowledgeBase(), [], [], output)
    wb = openpyxl.load_workbook(output)
    assert wb["Experience"].max_row == 1  # header-only, no data rows written
